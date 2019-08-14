# -*- coding: cp1251 -*-

import config
import text

import time
import datetime as dt
import random

from openpyxl import load_workbook 

def main_logic(txt_path, txt_xls_path, b_lang, txt_bp_name, txt_bp_link, b_name, b_decpt, b_ext_byp, list_do_opts, b_frame, b_cbArea, b_cbSLS, txt_DomainName, b_cbNamur, txt_Overrange, txt_Underrange, b_cbLF, txt_TripHys, b_cbAIBadIfLimited):
    out_file = file(txt_path, "w")

    full_data = getXLSData(txt_xls_path)
    mod_list = getModList(full_data)
    vtr_list = getVTRList(full_data)
    bp_ref_list = getBPRefList(full_data)
    c_areas_list = getUniqueAreas(full_data)
    c_sls_list = getSLSData(full_data)
    c_uniq_sls = getUniqueSLS(full_data)

    id = 1277825591
    id_vtr = 1355346828

    #===============================================================================
    # create Areas 
    #===============================================================================
    if b_cbArea:
        for area in c_areas_list:
            if area.upper() != text.defaultArea:
                writePlantAreaInfo(out_file, area.upper())

    #===============================================================================
    # create SLS 
    #===============================================================================
    if b_cbSLS:
        c_domain = text.SISDomainName
        
        if len(checkValue(txt_DomainName)) > 0:
            c_domain = txt_DomainName
        
        writeSISDomainInfo(out_file, c_domain)
        
        for sls in c_uniq_sls:
            #write SLS Header
            writeSLSHeader(out_file, sls.upper(), c_domain)
            
            #channels
            for i in xrange(1, 17):
                flag = False
                
                for sls_data in c_sls_list:
                    c_sls = checkValue(sls_data[0])
                    c_ch_dst = checkValue(sls_data[1])
                    c_ch_type = checkValue(sls_data[2])
                    c_ch = checkIntValue(sls_data[3])
                    c_ch_desc = checkValue(sls_data[4])
                                                                                
                    writeSLSChannels(out_file, c_ch_dst, c_ch, c_ch_type, c_ch_desc, c_ch_enab)
                    flag = True
                
                if not flag:                    
                    if i < 10:
                        c_ch = '0' + str(i)
                    else:
                        c_ch = str(i)
                        
                    c_ch_dst = sls.upper() + 'CH' + c_ch
                    c_ch_type = text.defaultChType
                    c_ch_desc = text.defaultChDesc
                    c_ch_enab = False
                    
                    writeSLSChannels(out_file, c_ch_dst, i, c_ch_type, c_ch_desc, c_ch_enab)
            
            #footer
            writeSLSFooter(out_file)
    
    #===============================================================================
    # fill text    
    #===============================================================================
    for c_mod_name in mod_list:       
        flag = False
                
        # positions
        c_pos_aidi = 1
        c_pos_do = 1
        
        # counters
        c_i_ai = 1
        c_i_di = 1
        c_i_do = 1
        c_i_avtr = 1
        c_i_dvtr = 1
        
        # lists
        con_list = []
        dst_list = []
        vtr_info_list = []
        vtr_desc_list = []        
        do_list = []
        scale_list = []
        byp_con_list = []
        ai_list = []
        
        c_bp_ref_diff = 0
        c_bp_ref_flag = False                    
        c_bp_flag = False
         
        for c in full_data:
            if c[7] == c_mod_name:
                
                c_SLS = checkValue(c[1])
                c_DST = checkValue(c[2])
                c_DSTType = checkValue(c[3])
                c_DSTCh = checkIntValue(c[4])
                c_DSTChDesc = checkValue(c[5])
                
                c_area = c[6]
                c_desc = checkValue(c[8])
                
                c_fb_name = checkValue(c[9])
                c_fb_type = checkValue(c[10])
                c_eu0, c_eu0_flag = checkFloatValue(c[11])
                c_eu100, c_eu100_flag = checkFloatValue(c[12])
                c_units = checkValue(c[13])
                c_decpt = checkIntValue(c[14])
                
                c_vtr_name = checkValue(c[15])
                c_vtr_type = checkValue(c[16])
                c_vtr_in = checkIntValue(c[17])                               
                c_vtr_num2trip = checkIntValue(c[18])
                c_vtr_det_type = checkValue(c[19])
                
                c_vtr_ptrip_lim, c_vtr_ptrip_flag = checkFloatValue(c[20])
                c_vtr_trip_lim, c_vtr_trip_flag = checkFloatValue(c[21])
                
                # check FB AI/DI/DO name
                if c_fb_name == '' and len(c_fb_type) > 0:
                    if c_fb_type == text.XLSSheetFBTypeList[0]:
                        c_fb_name = 'LSAI' + str(c_i_ai)
                        c_i_ai += 1
                    elif c_fb_type == text.XLSSheetFBTypeList[1]:
                        c_fb_name = 'LSDI' + str(c_i_di)
                        c_i_di += 1
                    else:
                        c_fb_name = 'LSDO' + str(c_i_do)
                        c_i_do += 1

                    if b_name and c_DST != '': 
                        c_fb_name = c_DST

                # check bypass name                
                c_byp_name = c_fb_name + '_BYP'
                
                # check FB VTR name
                if c_vtr_name == '' and len(c_vtr_type) > 0:
                    if c_vtr_type == text.XLSSheetVTRTypeList[0]:
                        c_vtr_name = 'LSAVTR' + str(c_i_avtr)
                        c_i_avtr += 1
                    elif c_vtr_type == text.XLSSheetVTRTypeList[1]:
                        c_vtr_name = 'LSDVTR' + str(c_i_dvtr)
                        c_i_dvtr += 1
                    
                    if b_name and c_DST != '': 
                        c_vtr_name = c_DST

                if c_vtr_type == text.XLSSheetVTRTypeList[0] and c_vtr_det_type == text.XLSSheetVTRDetTypeList[0]:
                    c_vtr_name += '_HH'
                elif c_vtr_type == text.XLSSheetVTRTypeList[0] and c_vtr_det_type == text.XLSSheetVTRDetTypeList[1]: 
                    c_vtr_name += '_LL'
                elif c_vtr_type == text.XLSSheetVTRTypeList[1] and c_vtr_name != 'LSDVTR' + str(c_i_dvtr -1):                           
                    c_vtr_name += '_DVTR'

                if (len(c_vtr_bypopts.split('8')) > 1 or c_bp_flag) and (c_fb_type == text.XLSSheetFBTypeList[0] or c_fb_type == text.XLSSheetFBTypeList[1]):
                    byp_flag = True
                    byp_con_list.append([c_byp_name, c_vtr_name, c_vtr_in])
                
                noi = getVTRNOI(vtr_list, c_mod_name, c_vtr_name)
                con_list.append([c_fb_name, c_fb_type, c_vtr_name, c_vtr_in, noi, c_bp_flag])
                dst_list.append([c_fb_name, c_fb_type, c_DST])
                
                if c_fb_type == text.XLSSheetFBTypeList[0]:
                    ai_list.append(c_fb_name)
                elif c_fb_type == text.XLSSheetFBTypeList[2]:
                    do_list.append(c_fb_name) 
                
                if c_fb_type == text.XLSSheetFBTypeList[0]:
                    scale_list.append([c_fb_name, c_vtr_name, c_vtr_in, c_eu0, c_eu0_flag, c_eu100, c_eu100_flag, c_units, c_decpt])
                
                if c_vtr_type == text.XLSSheetVTRTypeList[0] or c_vtr_type == text.XLSSheetVTRTypeList[1]:
                    vtr_info_list.append([c_vtr_name, c_vtr_ptrip_lim, c_vtr_ptrip_flag, c_vtr_trip_lim, c_vtr_trip_flag, c_vtr_in, c_vtr_type, c_vtr_det_type, c_vtr_num2trip])
                                       
                #===============================================================================
                # cycling write
                #===============================================================================
                # write info about AI/DI/DO                                     
                if flag and len(c_fb_type) > 0:
                    ai_scale_flag = False
                    
                    if  c_fb_type == text.XLSSheetFBTypeList[0]:
                        if c_vtr_in == 0 or c_vtr_in == 1:
                            ai_scale_flag = True
                        
                        writeLSAIInfo(out_file, c_fb_name, c_pos_aidi, id, c_bp_ref_diff, ai_scale_flag)
                    elif c_fb_type == text.XLSSheetFBTypeList[1]:
                        writeLSDIInfo(out_file, c_fb_name, c_pos_aidi, id, c_bp_ref_diff)
                    else:
                        writeLSDOInfo(out_file, c_fb_name, c_pos_do, id, c_bp_ref_diff)
               
                # write info about VTRs
                if flag and c_vtr_type == text.XLSSheetVTRTypeList[0]:
                    if c_vtr_in == 0 or c_vtr_in == 1:
                        if len(c_vtr_name) <= 16:
                            writeLSAVTRInfo(out_file, c_vtr_name, c_pos_aidi, id_vtr, noi, c_bp_ref_diff, c_bp_flag, byp_flag)
                        else:
                            writeLSAVTRInfo(out_file, 'LSAVTR' + str(c_i_avtr), c_pos_aidi, id_vtr, noi, c_bp_ref_diff, c_bp_flag, byp_flag)
                        
                elif flag and c_vtr_type == text.XLSSheetVTRTypeList[1]:
                    if c_vtr_in == 0 or c_vtr_in == 1:  
                        if len(c_vtr_name) <= 16:
                            writeLSDVTRInfo(out_file, c_vtr_name, c_pos_aidi, id_vtr, noi, c_bp_ref_diff, c_bp_flag, byp_flag)
                        else:
                            writeLSDVTRInfo(out_file, 'LSDVTR' + str(c_i_dvtr), c_pos_aidi, id_vtr, noi, c_bp_ref_diff, c_bp_flag, byp_flag)
                
                #===============================================================================
                # generate template       
                #===============================================================================
                if not flag:                                 
                    # line 1: header
                    out_file.write('SIF_MODULE TAG="' + c_mod_name.upper() + '" PLANT_AREA="' + c_area.upper() +'" CATEGORY=""\r\n')
                    
                    c_unix_time, c_date_time = getTime()
                    
                    # line 2: header
                    out_file.write(' user="administrator" time=' + c_unix_time + '/*' + c_date_time + '*/' + '\r\n')               
                
                    # line 3
                    out_file.write('{\r\n')
                    
                    # line 4: description
                    out_file.write('  DESCRIPTION="' + c_desc +'"\r\n')
                    # line 5: SLS
                    out_file.write('  LOGIC_SOLVER="' + c_SLS +'"\r\n')
                    
                    # header
                    out_file.write('  PRIMARY_CONTROL_DISPLAY=""\r\n')
                    out_file.write('  INSTRUMENT_AREA_DISPLAY="SIS_MOD_FP"\r\n')
                    out_file.write('  DETAIL_DISPLAY=""\r\n')
                    out_file.write('  TYPE=""\r\n')
                    out_file.write('  SUB_TYPE=""\r\n')

                    # write info about FB AI/DI/DO
                    ai_scale_flag = False
                    if  c_fb_type == text.XLSSheetFBTypeList[0]:                        
                        if c_vtr_in == 0 or c_vtr_in == 1:
                            ai_scale_flag = True
                                                    
                        writeLSAIInfo(out_file, c_fb_name, c_pos_aidi, id, c_bp_ref_diff, ai_scale_flag)
                    elif c_fb_type == text.XLSSheetFBTypeList[1]:
                        writeLSDIInfo(out_file, c_fb_name, c_pos_aidi, id, c_bp_ref_diff)
                    elif c_fb_type == text.XLSSheetFBTypeList[2]:
                        writeLSDOInfo(out_file, c_fb_name, c_pos_do, id, c_bp_ref_diff)
                            
                   # write info about voters:
                    if c_vtr_type == text.XLSSheetVTRTypeList[0]:
                        if c_vtr_in == 0 or c_vtr_in == 1: 
                            if len(c_vtr_name) <= 16:
                                writeLSAVTRInfo(out_file, c_vtr_name, c_pos_aidi, id_vtr, noi, c_bp_ref_diff, c_bp_flag, byp_flag)
                            else:
                                writeLSAVTRInfo(out_file, 'LSAVTR' + str(c_i_avtr), c_pos_aidi, id_vtr, noi, c_bp_ref_diff, c_bp_flag, byp_flag)
                            
                    elif c_vtr_type == text.XLSSheetVTRTypeList[1]:
                        if c_vtr_in == 0 or c_vtr_in == 1:
                            if len(c_vtr_name) <= 16:
                                writeLSDVTRInfo(out_file, c_vtr_name, c_pos_aidi, id_vtr, noi, c_bp_ref_diff, c_bp_flag, byp_flag)
                            else:
                                writeLSDVTRInfo(out_file, 'LSDVTR' + str(c_i_dvtr), c_pos_aidi, id_vtr, noi, c_bp_ref_diff, c_bp_flag, byp_flag) 
                                                                                       
                    flag = True   
                                
                # update counters                                     
                if  c_fb_type == text.XLSSheetFBTypeList[0] or c_fb_type == text.XLSSheetFBTypeList[1]:
                    c_pos_aidi += 1
                elif c_fb_type == text.XLSSheetFBTypeList[2]:
                    c_pos_do += 1

                id += 1
                id_vtr += 1

        out_file.write('  FBD_ALGORITHM\r\n')
        out_file.write('  {\r\n')
        
        # generate frame
        if b_frame:
            writeFrameBorder(out_file, b_lang)        
        
        # write info about connections
        writeConnectionsInfo(out_file, con_list, c_bp_name)
               
        out_file.write('  }\r\n')
        out_file.write('  ATTRIBUTE_INSTANCE NAME="VERSION"\r\n')
        out_file.write('  {\r\n')
        out_file.write('    VALUE { CV=1 }\r\n')
        out_file.write('  }\r\n')
        out_file.write('  ATTRIBUTE_INSTANCE NAME="EXEC_TIME"\r\n')
        out_file.write('  {\r\n')
        out_file.write('    VALUE { CV=0 }\r\n')
        out_file.write('  }\r\n')
        out_file.write('  ATTRIBUTE_INSTANCE NAME="SIF_ERRORS"\r\n')
        out_file.write('  {\r\n')
        out_file.write('    VALUE\r\n')
        out_file.write('    {\r\n')
        out_file.write('      ENUM_SET="$ls_sif_errors"\r\n')
        out_file.write('    }\r\n')
        out_file.write('  }\r\n')
        out_file.write('  ATTRIBUTE_INSTANCE NAME="SIF_ALERTS"\r\n')
        out_file.write('  {\r\n')
        out_file.write('    VALUE\r\n')
        out_file.write('    {\r\n')
        out_file.write('      ENUM_SET="$ls_sif_alerts"\r\n')
        out_file.write('    }\r\n')
        out_file.write('  }\r\n')
        out_file.write('  ATTRIBUTE_INSTANCE NAME="LS_NAME"\r\n')
        out_file.write('  {\r\n')
        out_file.write('    VALUE { CV="" }\r\n')
        out_file.write('  }\r\n')
        
        #write DSTs info
        writeDSTInfo(out_file, dst_list)

        # write voter params info
        writeVTRInfo(out_file, vtr_info_list, b_lang)
                
        # scale
        writeScaleInfo(out_file, scale_list, b_decpt)        

        # bypass permit ref        
        writeBPReference(out_file, c_bp_name, txt_bp_link)

        out_file.write('}\r\n')   

    out_file.close()

def writePlantAreaInfo(out_file, area):
    id = getIndex()
    c_unix_time, c_date_time = getTime()
                    
    out_file.write('PLANT_AREA NAME="' + area + '" INDEX=' + id + '\r\n')
    out_file.write(' user="administrator" time=' + c_unix_time + '/*' + c_date_time + '*/' + '\r\n')
    out_file.write('{\r\n')
    out_file.write('}\r\n')

def writeSISDomainInfo(out_file, domain):
    id = getIndex()
    c_unix_time, c_date_time = getTime()
    
    out_file.write('SISNET_DOMAIN NAME="' + domain + '" INDEX=1\r\n')
    out_file.write(' user="administrator" time=' + c_unix_time + '/*' + c_date_time + '*/' + '\r\n')
    out_file.write('{\r\n')
    out_file.write('}\r\n')

def getIndex():
    c_random = random.sample(['1', '2', '3', '4', '5', '6', '7', '8', '9', '0'],  2)
    
    id = ''    
    for c in c_random:
        id += c 

    return id


def writeSLSHeader(out_file, sls, c_domain):
    c_unix_time, c_date_time = getTime()
    
    out_file.write('LOGIC_SOLVER NAME="' + sls + '"\r\n')
    out_file.write(' user="administrator" time=' + c_unix_time + '/*' + c_date_time + '*/' + '\r\n')     
    out_file.write('{\r\n')
    out_file.write('  SISNET_DOMAIN="' + c_domain + '"\r\n')
    out_file.write('  LAST_GENERATED_CRC=0\r\n')
    out_file.write('  LAST_DOWNLOADED_CRC=0\r\n')
    out_file.write('  GLOBAL_SLOT=0\r\n')
    out_file.write('  DLOAD_SLOT=0\r\n')
    out_file.write('  PUBLISHER_TYPE=LOCAL\r\n')
    out_file.write('  PRIMARY_CONTROL_DISPLAY=""\r\n')
    out_file.write('  INSTRUMENT_AREA_DISPLAY="SIS_LSDEV_FP"\r\n')
    out_file.write('  DETAIL_DISPLAY=""\r\n')
    out_file.write('  SCAN_RATE=50_MILLISECONDS\r\n')
    out_file.write('  ENABLE_FAST_IO_UPDATES=F\r\n')
    out_file.write('  ASSIGNED_CARD=""\r\n')
    out_file.write('  TEST_INTERVAL_TIME=0\r\n')
    out_file.write('  REMINDER_TIME=0\r\n')
    out_file.write('  AUTO_PROOF_TEST=F\r\n')
    out_file.write('  REDUCED_STATUS=F\r\n')

def writeSLSChannels(out_file, c_ch_dst, c_ch, c_ch_type, c_ch_desc, c_ch_enab):    
    out_file.write('  SIMPLE_IO_CHANNEL POSITION=' + str(c_ch) + ' DEFINITION="' + c_ch_type + '"\r\n')
    out_file.write('  {\r\n')
    out_file.write('    DESCRIPTION="' + c_ch_desc + '"\r\n')
    out_file.write('    ENABLED=' + str(c_ch_enab)[0:1] + '\r\n')
    out_file.write('    DEVICE_SIGNAL_TAG="' + c_ch_dst + '"\r\n')
    out_file.write('    HART_LONG_TAG=""\r\n')
    out_file.write('  }\r\n')

def writeSLSFooter(out_file):
    out_file.write('}\r\n')

def writeLSAIInfo(out_file, c_fb_name, i, id, diff, flag):
    out_file.write('  FUNCTION_BLOCK NAME="' + c_fb_name + '" DEFINITION="LSAI"\r\n')
    out_file.write('  {\r\n')
    out_file.write('    DESCRIPTION="Analog Input"\r\n')
    out_file.write('    ID=' + str(id) + '\r\n')
    out_file.write('    RECTANGLE= { X=50 Y=' + str(100 + (i - 1)*150 + diff) +' H=80 W=140 }\r\n')
    
    if flag: 
        out_file.write('    ADDITIONAL_CONNECTOR NAME="OUT_SCALE" TYPE=OUTPUT { ATTRIBUTE="OUT_SCALE" }\r\n')
    
    out_file.write('  }\r\n')

def writeLSDIInfo(out_file, c_fb_name, i, id, diff):
    out_file.write('  FUNCTION_BLOCK NAME="' + c_fb_name + '" DEFINITION="LSDI"\r\n')
    out_file.write('  {\r\n')
    out_file.write('    DESCRIPTION="Discrete Input"\r\n')
    out_file.write('    ID=' + str(id) + '\r\n')
    out_file.write('    RECTANGLE= { X=50 Y=' + str(100 + (i - 1)*150 + diff) +' H=80 W=140 }\r\n')
    out_file.write('  }\r\n')

def writeLSDOInfo(out_file, c_fb_name, i, id, diff):
    out_file.write('  FUNCTION_BLOCK NAME="' + c_fb_name + '" DEFINITION="LSDO"\r\n')
    out_file.write('  {\r\n')
    out_file.write('    DESCRIPTION="Discrete Output"\r\n')
    out_file.write('    ID=' + str(id) + '\r\n')
    out_file.write('    RECTANGLE= { X=1110 Y=' + str(100 + (i - 1)*150 + diff) +' H=80 W=140 }\r\n')
    out_file.write('  }\r\n')

def writeLSAVTRInfo(out_file, c_vtr_name, i, id, noi, diff):
    out_file.write('  FUNCTION_BLOCK NAME="' + c_vtr_name + '" DEFINITION="LSAVTR"\r\n')
    out_file.write('  {\r\n')
    out_file.write('    DESCRIPTION="Analog Voter"\r\n')
    out_file.write('    ID=' + str(id) + '\r\n')
    out_file.write('    RECTANGLE= { X=400 Y=' + str(100 + (i - 1)*150 + diff) + ' H=80 W=140 }\r\n')
    out_file.write('    EXTENSIBLE_ATTRIBUTE { NAME="TRIP_VOTE_IN"  COUNT=' + str(noi) + ' }\r\n')
    out_file.write('    EXTENSIBLE_ATTRIBUTE { NAME="PRE_VOTE_IN"  COUNT=' + str(noi) + ' }\r\n')
    out_file.write('    EXTENSIBLE_ATTRIBUTE { NAME="IN"  COUNT=' + str(noi) + ' }\r\n')
    out_file.write('    EXTENSIBLE_ATTRIBUTE { NAME="DESC"  COUNT=' + str(noi) + ' }\r\n')
    out_file.write('    EXTENSIBLE_ATTRIBUTE { NAME="BYPASS"  COUNT=' + str(noi) + ' }\r\n')
    out_file.write('    ADDITIONAL_CONNECTOR NAME="IN_SCALE" TYPE=INPUT { ATTRIBUTE="IN_SCALE" }\r\n')
    out_file.write('  }\r\n')

def writeLSDVTRInfo(out_file, c_vtr_name, i, id, noi, diff):
    out_file.write('  FUNCTION_BLOCK NAME="' + c_vtr_name + '" DEFINITION="LSDVTR"\r\n')
    out_file.write('  {\r\n')
    out_file.write('    DESCRIPTION="Discrete Voter"\r\n')
    out_file.write('    ID=' + str(id) + '\r\n')
    out_file.write('    RECTANGLE= { X=400 Y=' + str(100 + (i - 1)*150 + diff) + ' H=80 W=140 }\r\n')
    out_file.write('    EXTENSIBLE_ATTRIBUTE { NAME="TRIP_VOTE_IN"  COUNT=' + str(noi) + ' }\r\n')
    out_file.write('    EXTENSIBLE_ATTRIBUTE { NAME="IN_D"  COUNT=' + str(noi) + ' }\r\n')
    out_file.write('    EXTENSIBLE_ATTRIBUTE { NAME="DESC"  COUNT=' + str(noi) + ' }\r\n')
    out_file.write('    EXTENSIBLE_ATTRIBUTE { NAME="BYPASS"  COUNT=' + str(noi) + ' }\r\n')

def writeConnectionsInfo(out_file, con_list, bp_name):
    for c in con_list:
        c_fb_name = c[0]
        c_fb_type = c[1]
        c_vtr_name = c[2]
        c_vtr_in = c[3]
        c_noi = c[4]
        c_bp_flag = c[5]
               
        # ai
        if c_fb_type == text.XLSSheetFBTypeList[0]:
            out_file.write('    WIRE SOURCE="' + c_fb_name + '/OUT" DESTINATION="' + c_vtr_name + '/IN' + str(c_vtr_in) + '" { IS_FEEDBACK_WIRE=T SEGMENT { INDEX=2 ORIENTATION=VERTICAL ORDINATE=' + str(200 + c_vtr_in*10) + ' } }\r\n')
            out_file.write('    WIRE SOURCE="' + c_fb_name + '/OUT_SCALE" DESTINATION="' + c_vtr_name + '/IN_SCALE" { SEGMENT { INDEX=2 ORIENTATION=VERTICAL ORDINATE=210 } }\r\n') 
        
        # di
        elif c_fb_type == text.XLSSheetFBTypeList[1]:
            out_file.write('    WIRE SOURCE="' + c_fb_name + '/OUT_D" DESTINATION="' + c_vtr_name + '/IN_D' + str(c_vtr_in) + '" { IS_FEEDBACK_WIRE=T SEGMENT { INDEX=2 ORIENTATION=VERTICAL ORDINATE=' + str(200 + c_vtr_in*10) + ' } }\r\n')        
        
        # bypass permit
        if c_bp_flag and (c_fb_type == text.XLSSheetFBTypeList[0] or c_fb_type == text.XLSSheetFBTypeList[1]):
            out_file.write('    WIRE SOURCE="' + bp_name + '" DESTINATION="' + c_vtr_name + '/BYPASS_PERMIT" { SEGMENT { INDEX=2 ORIENTATION=VERTICAL ORDINATE=380 } }\r\n')

def writeFrameBorder(out_file, b_lang):
    name = randomizeIt(8)
    name += '-'
    name += randomizeIt(4)
    name += '-'
    name += randomizeIt(4)
    name += '-'
    name += randomizeIt(4)
    name += '-'

    txt_rev = text.FHXRev[0]
    txt_date = text.FHXDate[0]
    txt_author = text.FHXAuthor[0]
    txt_comments = text.FHXComments[0]
    
    out_file.write('    GRAPHICS ALGORITHM=FBD\r\n')
    out_file.write('    {\r\n')
    out_file.write('      BOX_GRAPHIC\r\n')
    out_file.write('      {\r\n')
    out_file.write('        NAME="{' + name + randomizeIt(12) + '}"\r\n')
    out_file.write('        RECTANGLE= { X=5 Y=5 H=1990 W=1290 }\r\n')
    out_file.write('        LINE_STYLE=SOLID\r\n')
    out_file.write('        LINE_WIDTH=1\r\n')
    out_file.write('        LINE_COLOR= { RED=0 GREEN=0 BLUE=0 }\r\n')
    out_file.write('        FGD_COLOR= { RED=255 GREEN=255 BLUE=255 }\r\n')
    out_file.write('        BGD_COLOR= { RED=0 GREEN=0 BLUE=0 }\r\n')
    out_file.write('        FILL_PATTERN=""\r\n')
    out_file.write('      }\r\n')
    out_file.write('      BOX_GRAPHIC\r\n')
    out_file.write('      {\r\n')
    out_file.write('        NAME="{' + name + randomizeIt(12) + '}"\r\n')
    out_file.write('        RECTANGLE= { X=5 Y=1800 H=30 W=1290 }\r\n')
    out_file.write('        LINE_STYLE=SOLID\r\n')
    out_file.write('        LINE_WIDTH=1\r\n')
    out_file.write('        LINE_COLOR= { RED=0 GREEN=0 BLUE=0 }\r\n')
    out_file.write('        FGD_COLOR= { RED=255 GREEN=255 BLUE=128 }\r\n')
    out_file.write('        BGD_COLOR= { RED=0 GREEN=0 BLUE=0 }\r\n')
    out_file.write('        FILL_PATTERN=""\r\n')
    out_file.write('      }\r\n')

    out_file.write('      TEXT_GRAPHIC\r\n')
    out_file.write('      {\r\n')
    out_file.write('        NAME="{' + name + randomizeIt(12) + '}"\r\n')
    out_file.write('        ORIGIN= { X=35 Y=1840 }\r\n')
    out_file.write('        END= { X=65 Y=1868 }\r\n')
    out_file.write('        TEXT="' + txt_rev + ':\r\n')
    out_file.write('------"\r\n')
    out_file.write('      }\r\n')
    out_file.write('      TEXT_GRAPHIC\r\n')
    out_file.write('      {\r\n')
    out_file.write('        NAME="{' + name + randomizeIt(12) + '}"\r\n')
    out_file.write('        ORIGIN= { X=115 Y=1840 }\r\n')
    out_file.write('        END= { X=148 Y=1868 }\r\n')
    out_file.write('        TEXT="' + txt_date + ':\r\n')
    out_file.write('------"\r\n')
    out_file.write('      }\r\n')
    out_file.write('      TEXT_GRAPHIC\r\n')
    out_file.write('      {\r\n')
    out_file.write('        NAME="{' + name + randomizeIt(12) + '}"\r\n')
    out_file.write('        ORIGIN= { X=198 Y=1840 }\r\n')
    out_file.write('        END= { X=242 Y=1868 }\r\n')
    out_file.write('        TEXT="' + txt_author + ':\r\n')
    out_file.write('---------"\r\n')
    out_file.write('      }\r\n')
    out_file.write('      TEXT_GRAPHIC\r\n')
    out_file.write('      {\r\n')
    out_file.write('        NAME="{' + name + randomizeIt(12) + '}"\r\n')
    out_file.write('        ORIGIN= { X=292 Y=1840 }\r\n')
    out_file.write('        END= { X=353 Y=1868 }\r\n')
    out_file.write('        TEXT="' + txt_comments + ':\r\n')
    out_file.write('-------------"\r\n')
    out_file.write('      }\r\n')
    
    out_file.write('    }\r\n')

def randomizeIt(noi):
    c_random = random.sample(['A', 'B', 'C', 'D', 'E', 'F', '1', '2', '3', '4', '5', '6', '7', '8', '9', '0'],  noi)
    
    cstr = ''
    
    for c in c_random:
        cstr += c 

    return cstr

def writeDSTInfo(out_file, dst_list):
    for c in dst_list:
        c_fb_name = c[0]
        c_fb_type = c[1]
        c_DST = c[2]
    
        if c_fb_type == text.XLSSheetFBTypeList[0]:
            out_file.write('  ATTRIBUTE_INSTANCE NAME="' + c_fb_name + '/IO_IN"\r\n')
            out_file.write('  {\r\n')
            out_file.write('    VALUE { REF="//' + c_DST + '/FIELD_VAL_PCT" CLASS=FLOAT_INPUT }\r\n')
            out_file.write('  }\r\n')
        elif c_fb_type == text.XLSSheetFBTypeList[1]:
            out_file.write('  ATTRIBUTE_INSTANCE NAME="' + c_fb_name + '/IO_IN"\r\n')
            out_file.write('  {\r\n')
            out_file.write('    VALUE { REF="//' + c_DST + '/FIELD_VAL_D" CLASS=DISCRETE_INPUT }\r\n')
            out_file.write('  }\r\n')
        elif c_fb_type == text.XLSSheetFBTypeList[2]:
            out_file.write('  ATTRIBUTE_INSTANCE NAME="' + c_fb_name + '/IO_OUT"\r\n')
            out_file.write('  {\r\n')
            out_file.write('    VALUE { REF="//' + c_DST + '/OUT_D" CLASS=DISCRETE_OUTPUT }\r\n')
            out_file.write('  }\r\n')

def writeVTRInfo(out_file, vtr_info_list, b_lang):  
    for c in vtr_info_list:
        c_vtr_name = c[0]
        c_vtr_ptrip_lim = c[1] 
        c_vtr_ptrip_flag = c[2]
        c_vtr_trip_lim = c[3]
        c_vtr_trip_flag = c[4]
        c_vtr_in = c[5]
        c_vtr_type = c[6]
        c_vtr_det_type = c[7]
        c_vtr_num2trip = c[8]        
        c_vtr_st_opts = c[9]
        
        ptrip = c_vtr_ptrip_lim
        trip = c_vtr_trip_lim
        
        # pre trip lim
        if (c_vtr_in == 0 or c_vtr_in == 1) and c_vtr_type == text.XLSSheetVTRTypeList[0]:
            out_file.write('  ATTRIBUTE_INSTANCE NAME="' + c_vtr_name + '/PRE_TRIP_LIM"\r\n')
            out_file.write('  {\r\n')
            out_file.write('    VALUE { CV=' + str(ptrip) + ' }\r\n')
            out_file.write('  }\r\n')
        
        # trip lim
        if (c_vtr_in == 0 or c_vtr_in == 1) and c_vtr_type == text.XLSSheetVTRTypeList[0]:
            out_file.write('  ATTRIBUTE_INSTANCE NAME="' + c_vtr_name + '/TRIP_LIM"\r\n')
            out_file.write('  {\r\n')
            out_file.write('    VALUE { CV=' + str(trip) + ' }\r\n')
            out_file.write('  }\r\n')
        
        # num to trip
        if c_vtr_in == 0 or c_vtr_in == 1:
            if c_vtr_num2trip == 0:
                c_vtr_num2trip = 1
            
            out_file.write('  ATTRIBUTE_INSTANCE NAME="' + c_vtr_name + '/NUM_TO_TRIP"\r\n')
            out_file.write('  {\r\n')
            out_file.write('    VALUE { CV=' + str(c_vtr_num2trip) + ' }\r\n')
            out_file.write('  }\r\n')
        
        # detect type
        if (c_vtr_in == 0 or c_vtr_in == 1) and c_vtr_type == text.XLSSheetVTRTypeList[0]:
            det_type = text.XLSSheetVTRDetTypeList[0]
            
            if b_lang:
                det_type = text.XLSSheetVTRDetTypeListRus[0]
                
                if c_vtr_det_type == text.XLSSheetVTRDetTypeList[0]:
                    det_type = text.XLSSheetVTRDetTypeListRus[0]
                elif c_vtr_det_type == text.XLSSheetVTRDetTypeList[1]:
                    det_type = text.XLSSheetVTRDetTypeListRus[1]
            else:
                if c_vtr_det_type == text.XLSSheetVTRDetTypeList[0]:
                    det_type = text.XLSSheetVTRDetTypeList[0]
                elif c_vtr_det_type == text.XLSSheetVTRDetTypeList[1]:
                    det_type = text.XLSSheetVTRDetTypeList[1]                
                    
        
            out_file.write('  ATTRIBUTE_INSTANCE NAME="' + c_vtr_name + '/DETECT_TYPE"\r\n')
            out_file.write('  {\r\n')
            out_file.write('    VALUE\r\n')
            out_file.write('    {\r\n')
            out_file.write('      SET="$detect_type"\r\n')
            out_file.write('      STRING_VALUE="' + det_type + '"\r\n')
            out_file.write('      CHANGEABLE=F\r\n')
            out_file.write('    }\r\n')
            out_file.write('  }\r\n')

        # status opts
        if (c_vtr_in == 0 or c_vtr_in == 1):
            st_opts = text.XLSSheetVTRStOptsList[1]
            
            if b_lang:
                st_opts = text.XLSSheetVTRDetTypeListRus[0]
                
                if c_vtr_st_opts == text.XLSSheetVTRStOptsList[1]:
                    st_opts = text.XLSSheetVTRStOptsListRus[0]
                elif c_vtr_st_opts == text.XLSSheetVTRStOptsList[2]:
                    st_opts = text.XLSSheetVTRStOptsListRus[1]
                elif c_vtr_st_opts == text.XLSSheetVTRStOptsList[3]:
                    st_opts = text.XLSSheetVTRStOptsListRus[2]                    
            else:
                if c_vtr_st_opts == text.XLSSheetVTRStOptsList[1]:
                    st_opts = text.XLSSheetVTRStOptsList[0]
                elif c_vtr_st_opts == text.XLSSheetVTRStOptsList[2]:
                    st_opts = text.XLSSheetVTRStOptsList[1] 
                elif c_vtr_st_opts == text.XLSSheetVTRStOptsList[3]:
                    st_opts = text.XLSSheetVTRStOptsList[2]

            out_file.write('  ATTRIBUTE_INSTANCE NAME="' + c_vtr_name + '/STATUS_OPT"\r\n')
            out_file.write('  {\r\n')
            out_file.write('    VALUE\r\n')
            out_file.write('    {\r\n')
            out_file.write('      SET="$votr_status_opt"\r\n')
            out_file.write('      STRING_VALUE="' + st_opts + '"\r\n')
            out_file.write('      CHANGEABLE=F\r\n')
            out_file.write('    }\r\n')
            out_file.write('  }\r\n')

def writeScaleInfo(out_file, scale_list, b_decpt):
    for c in scale_list:
        c_fb_name = c[0]
        c_vtr_name = c[1]
        c_vtr_in = c[2]
        c_eu0 = c[3]
        c_eu0_flag = c[4]
        c_eu100 = c[5]
        c_eu100_flag = c[6]
        c_units = c[7]
        c_decpt = c[8]
        
        if c_eu0_flag == False:
            c_eu0 = 0
        
        if c_eu100_flag == False:
            c_eu100 = 100
        
        if b_decpt and c_decpt == 0:
            if c_eu100 < 1:
                c_decpt = 3
            elif c_eu100 < 100:
                c_decpt = 2
            elif c_eu100 < 1000:
                c_decpt = 1
            else:
                c_decpt = 0
        
        out_file.write('  ATTRIBUTE_INSTANCE NAME="' + c_fb_name + '/OUT_SCALE"\r\n')
        out_file.write('  {\r\n')
        out_file.write('    VALUE { EU100=' + str(c_eu100) + ' EU0=' + str(c_eu0) + ' UNITS="' + c_units + '" DECPT=' + str(c_decpt) + ' }\r\n')
        out_file.write('  }\r\n')
        
        if c_vtr_in == 0 or c_vtr_in == 1:
            out_file.write('  ATTRIBUTE_INSTANCE NAME="' + c_vtr_name + '/IN_SCALE"\r\n')
            out_file.write('  {\r\n')
            out_file.write('    VALUE { EU100=' + str(c_eu100) + ' EU0=' + str(c_eu0) + ' UNITS="' + c_units + '" DECPT=' + str(c_decpt) + ' }\r\n')
            out_file.write('  }\r\n')

def writeBPReference(out_file, bp_name, bp_ref):
    if len(bp_ref) > 0:
        out_file.write('  ATTRIBUTE_INSTANCE NAME="' + bp_name + '"\r\n')
        out_file.write('  {\r\n')
        out_file.write('    VALUE { REF="//' + bp_ref + '" }\r\n')
        out_file.write('  }\r\n')

def checkValue(val):
    out_str = ''
    
    if val != None:
        out_str = val
    
    return out_str

def checkIntValue(val):
    out_int = 0
    
    if val != None:
        out_int = int(val)
    
    return out_int

def checkFloatValue(val):
    flag = False
    
    out_float = 0
    
    if val != None:
        out_float = float(val)
        flag = True
    
    return out_float, flag

def getTime():
    c_unix_time = str(time.time()).split('.')[0]
    
    d = dt.datetime.now()
    c_month = d.strftime("%B")[0:3]

    c_datetime = d.strftime("%d") + '-' + c_month.decode('cp1251') + d.strftime("-%Y %H:%M:%S")
    
    return c_unix_time, c_datetime
    

def getXLSData(c_xls_path):
    wb = load_workbook(filename = c_xls_path, use_iterators = True)

    ws = wb.get_sheet_by_name(name = 'Template')
        
    data = []
    
    for row in ws.iter_rows(): # it brings a new method: iter_rows()
        tmp_data = []
    
        for cell in row:
            if cell.row > 2:
                tmp_data.append(cell.internal_value)
    
        if len(tmp_data) > 0:
            data.append(tmp_data)
    
    return data

def getUniqueAreas(full_data):
    areas_list = []

    for c in full_data:
        flag = False
         
        for x in areas_list:
            if x == c[6]:
                flag = True
        
        if not flag:
            areas_list.append(c[6])
   
    return areas_list

def getUniqueSLS(full_data):
    u_sls_list = []

    for c in full_data:
        flag = False
         
        for x in u_sls_list:
            if x == c[1]:
                flag = True
        
        if not flag:
            u_sls_list.append(c[1])
   
    return u_sls_list

def getSLSData(full_data):
    sls_list = []
    
    for c in full_data:
        sls_list.append([c[1], c[2], c[3], checkIntValue(c[4]), c[5]])
    
    return sls_list

def getModList(full_data):
    mod_list = []
    
    for c in full_data:
        flag = False
         
        for x in mod_list:
            if x == c[7]:
                flag = True
        
        if not flag:
            mod_list.append(c[7])
            
    return mod_list

def getVTRList(full_data):
    tmp_vtr_list = []
     
    for c in full_data:
        c_mod_name = checkValue(c[7])
        c_vtr_name = checkValue(c[15])
        c_vtr_in = checkValue(c[17])
        
        if checkValue(c[24]) == 'Yes':
            c_vtr_bp_flag =  True
        else:
            c_vtr_bp_flag =  False

        tmp_vtr_list.append([c_mod_name, c_vtr_name, c_vtr_in])
    
    tmp_vtr_list2 = tmp_vtr_list
    vtr_list = []
    
    for c in tmp_vtr_list:
        c_mod_name = ''
        c_vtr_name = ''
        c_vtr_in = 1

        for x in tmp_vtr_list2:
            if (c[0] == x[0] and c[0] != '') and (c[1] == x[1] and c[1] != ''): 
                if int(x[2]) > int(c[2]):
                    c_mod_name = c[0] 
                    c_vtr_name = c[1]
                    c_vtr_in = int(x[2])
        
        if c_vtr_in > 1:
            flag = True
            
            for q in vtr_list:
                if q[0] == c_mod_name and q[1] == c_vtr_name and q[2] == c_vtr_in:
                    flag = False
            
            if flag:
                vtr_list.append([c_mod_name, c_vtr_name, c_vtr_in])

    return vtr_list

def getVTRNOI(vtr_list, c_mod_name, c_vtr_name):
    noi = 1
    
    if c_vtr_name[len(c_vtr_name) - 3: len(c_vtr_name)] == '_HH' or c_vtr_name[len(c_vtr_name) - 3: len(c_vtr_name)] == '_LL':
        c_vtr_name = c_vtr_name[0:len(c_vtr_name) - 3]
    elif c_vtr_name[len(c_vtr_name) - 5: len(c_vtr_name)] == '_DVTR':
        c_vtr_name = c_vtr_name[0:len(c_vtr_name) - 5]

    for c in vtr_list:
        if c[0] == c_mod_name and c[1] == c_vtr_name:
            noi = int(c[2])

    return noi

def getBPRefList(full_data):
    bp_ref_list = []
    mod_list = getModList(full_data)
    
    for c in mod_list:
        flag = False
        
        for x in full_data:
            if x[7] == c and checkValue(x[24]) == 'Yes':
                flag = True
                break
        
        bp_ref_list.append([c, flag])

    return bp_ref_list